"""Generate a print-ready Weekly Strategic Motivation Tracker (A4 landscape)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

OUTPUT = r"C:\Users\HP\Documents\PROJECTS\slide-desk\Weekly_Motivation_Tracker.xlsx"

# Managers — get the ★ MGR badge in column B
MANAGERS = {
    "Namugga Martha",
    "Masembe Sendi Joseph",
    "Muzuva Joshua",
    "Fred Sseginda",
}

# Team -> light tint for the name cell, members
groups = [
    ("Focust", "FFE8D6", [
        "Masembe Sendi Joseph",
        "Wamani Joshua",
        "Muwanguzi Alvin Kiggundu",
        "Timon W. Mesulam",
        "Yapyeko Rebecca",
        "Kiire Constantine",
        "Ogenrwoth Jim Frank",
    ]),
    ("Create Project", "E2F0CB", [
        "Muzuva Joshua",
        "Turyasiima Crispus",
        "Joshua Prince Aita",
        "Lauben Mpairwe",
        "Otai Joseph",
    ]),
    ("EMS", "DDEAF6", [
        "Aheebwa Steven",
        "Fred Sseginda",
        "Konso Vanessa Rebecca",
        "Aule Ian",
    ]),
    ("GA", "FFF1B8", [
        "Masaba Shamsa",
        "Namugga Martha",
        "Ahwera Shifrah",
        "Namutebi Sulaina Luwambya",
    ]),
    ("AI Video Platform", "F4D9E8", [
        "Kisakye Martha Janepher",
        "Salimu Kabogere",
    ]),
    ("Matsumoto San's Project", "E8DFF5", [
        "Muwanguzi Jovic Biralo",
    ]),
    ("Urban Flooding", "D6F0F0", [
        "Tusiime Mark",
    ]),
    ("AI Super Agent", "F8D7DA", [
        "Shakiran Nannyombi",
    ]),
]

wb = Workbook()
ws = wb.active
ws.title = "Motivation Tracker"

# ---------- Style primitives ----------
thin = Side(border_style="thin", color="555555")
medium = Side(border_style="medium", color="222222")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_strong = Border(left=medium, right=medium, top=medium, bottom=medium)

NAVY = "1F2F4D"
GOLD = "F5C242"
LIGHT_BG = "F5F7FB"
LEGEND_BG = "EEF2F8"
FOOTER_BG = "FFF3B8"

title_font = Font(name="Calibri", size=20, bold=True, color=NAVY)
subtitle_font = Font(name="Calibri", size=11, italic=True, color="555555")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
field_label_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
field_value_font = Font(name="Calibri", size=11, color=NAVY)
legend_label_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
legend_item_font = Font(name="Calibri", size=10, color=NAVY)
mgr_font = Font(name="Calibri", size=10, bold=True, color="6B4A00")
nonmgr_font = Font(name="Calibri", size=10, color="AAAAAA")

header_fill = PatternFill("solid", fgColor=NAVY)
field_label_fill = PatternFill("solid", fgColor=NAVY)
field_value_fill = PatternFill("solid", fgColor="FFFFFF")
legend_label_fill = PatternFill("solid", fgColor=NAVY)
legend_body_fill = PatternFill("solid", fgColor=LEGEND_BG)
mgr_fill = PatternFill("solid", fgColor=GOLD)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)


def box(cell, fill=None, border=border_all, align=center, font=None):
    if fill is not None:
        cell.fill = fill
    cell.border = border
    cell.alignment = align
    if font is not None:
        cell.font = font


# ---------- Row 1: Title ----------
ws.merge_cells("A1:I1")
ws["A1"] = "WEEKLY STRATEGIC MOTIVATION TRACKER"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34

# ---------- Row 2: Subtitle ----------
ws.merge_cells("A2:I2")
ws["A2"] = "Spot high-impact contributors  •  Time recognition deliberately  •  Keep no one invisible"
ws["A2"].font = subtitle_font
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# ---------- Row 3: Week of / Team Lead — bold visible boxes ----------
# A3 label "WEEK OF" navy fill white bold
# B3:D3 merged white box with thick border
# E3 label "TEAM LEAD" navy fill white bold
# F3:I3 merged white box with thick border
ws["A3"] = "  WEEK OF"
box(ws["A3"], fill=field_label_fill, border=border_strong,
    align=Alignment(horizontal="left", vertical="center"),
    font=field_label_font)

ws.merge_cells("B3:D3")
box(ws["B3"], fill=field_value_fill, border=border_strong,
    align=Alignment(horizontal="left", vertical="center", indent=1),
    font=field_value_font)
# also border the merged-out cells for consistent print appearance
for col_letter in ["C", "D"]:
    ws[f"{col_letter}3"].border = border_strong

ws["E3"] = "  TEAM LEAD"
box(ws["E3"], fill=field_label_fill, border=border_strong,
    align=Alignment(horizontal="left", vertical="center"),
    font=field_label_font)

ws.merge_cells("F3:I3")
box(ws["F3"], fill=field_value_fill, border=border_strong,
    align=Alignment(horizontal="left", vertical="center", indent=1),
    font=field_value_font)
for col_letter in ["G", "H", "I"]:
    ws[f"{col_letter}3"].border = border_strong

ws.row_dimensions[3].height = 30

# ---------- Row 4: spacer ----------
ws.row_dimensions[4].height = 6

# ---------- Row 5: Legend ----------
ws["A5"] = "  LEGEND"
box(ws["A5"], fill=legend_label_fill, border=border_strong,
    align=Alignment(horizontal="left", vertical="center"),
    font=legend_label_font)

ws.merge_cells("B5:I5")
ws["B5"] = ("⚡ High Impact     "
            "💬 Touchpoint sent     "
            "🌟 Public shout-out     "
            "🔥 At-risk / Burnout watch     "
            "🛟 Blocker — needs support")
box(ws["B5"], fill=legend_body_fill, border=border_strong,
    align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    font=legend_item_font)
for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
    ws[f"{col_letter}5"].border = border_strong

ws.row_dimensions[5].height = 28

# ---------- Row 6: spacer ----------
ws.row_dimensions[6].height = 6

# ---------- Row 7: Table header ----------
header_row = 7
headers = ["Employee  (Team)", "Role", "Mon", "Tue", "Wed", "Thu", "Fri",
           "Weekly Impact", "Action Taken / Notes"]
for col_idx, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border_strong
ws.row_dimensions[header_row].height = 28

# ---------- Data rows, grouped by team ----------
name_inline_bold = InlineFont(rFont="Calibri", sz=11, b=True, color=NAVY)
team_inline = InlineFont(rFont="Calibri", sz=9, i=True, color="666666")

row = header_row + 1
for team_name, color, members in groups:
    name_fill = PatternFill("solid", fgColor=color)
    for member in members:
        # Column A — Name + (Team) on two lines
        rich = CellRichText(
            TextBlock(name_inline_bold, member),
            TextBlock(team_inline, f"\n({team_name})"),
        )
        a = ws.cell(row=row, column=1)
        a.value = rich
        a.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
        a.fill = name_fill
        a.border = border_all

        # Column B — Manager badge
        b = ws.cell(row=row, column=2)
        if member in MANAGERS:
            b.value = "★ MGR"
            b.font = mgr_font
            b.fill = mgr_fill
        else:
            b.value = "—"
            b.font = nonmgr_font
        b.alignment = center
        b.border = border_all

        # Mon–Fri + Weekly Impact + Notes
        for col in range(3, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = border_all
            cell.alignment = center if col < 9 else left
            cell.font = Font(name="Calibri", size=10)

        ws.row_dimensions[row].height = 36
        row += 1

last_row = row - 1

# ---------- Footer / Thursday Audit reminder ----------
ws.row_dimensions[row].height = 6  # spacer
footer_row = row + 1
ws.merge_cells(start_row=footer_row, start_column=1,
               end_row=footer_row, end_column=9)
fcell = ws.cell(row=footer_row, column=1,
                value=("THURSDAY AUDIT  →  Scan for ⚡ rows with no 💬 yet. "
                       "Those are the people to recognise before Friday closes."))
fcell.font = Font(name="Calibri", size=11, bold=True, italic=True, color=NAVY)
fcell.fill = PatternFill("solid", fgColor=FOOTER_BG)
fcell.alignment = Alignment(horizontal="center", vertical="center")
for col in range(1, 10):
    ws.cell(row=footer_row, column=col).border = border_strong
ws.row_dimensions[footer_row].height = 30

# ---------- Column widths tuned for A4 landscape ----------
widths = {
    "A": 30,   # Name + (Team)
    "B": 9,    # Role badge
    "C": 7, "D": 7, "E": 7, "F": 7, "G": 7,  # Mon–Fri
    "H": 14,   # Weekly Impact
    "I": 32,   # Action Taken
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ---------- Print setup ----------
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_options.horizontalCentered = True
ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                              header=0.2, footer=0.2)
ws.print_title_rows = f"{header_row}:{header_row}"
ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Rows: {last_row - header_row} members across {len(groups)} teams")
print(f"Managers flagged: {sorted(MANAGERS)}")
